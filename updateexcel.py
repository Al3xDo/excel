import openpyxl
# xu ly du lieu dau vao
exp={}
while True:
    print('Nhập mặt hàng cần chỉnh sửa giá: (nhập space để dừng)')
    name=input()
    if name==' ':
        break
    else:
        print('Nhập giá tiền:')
        exp[name]=input()
# xu ly dia chi excel
'''while True:
    print('Nhập địa chỉ file excel cần so (tên hàng hóa,...) (viết hoa ô nhớ):')
    dc1=input()
    print('Nhập địa chỉ file excel cần update ( viết hoa)')
    dc2=input()'''
# xu ly excel
print('Đang upload file excel...')
wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb.get_sheet_by_name('Sheet')
sheet.freeze_panes='A2'
# xu ly 2
print('Bắt đầu cập nhật các giá trị đã nhập:')
for obj in sheet['A2':'D23758']:
    for k,v in exp.items():
        if obj[0].value==k:
            obj[1].value=float(v)
            obj[3].value=float(v)*float(obj[2].value)
# xu ly cách 2
'''for rowN in range(2,sheet.get_highest_row()):
    name=sheet.cell(row=rowN, colum=1 ''''''= cột A'''''').value
    if name in exp:
        sheet.cell(row=rowN,colum=2).value=exp[name]'''
wb.save('produces1sale.xlsx')
print('Đã hoàn thành')

