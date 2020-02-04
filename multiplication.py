import openpyxl
from openpyxl.styles import Font
# xu ly file excel
wb=openpyxl.load_workbook('text.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
print('Nhập n')
n=int(input())
# tao font chu
'''boldf=font(size='12',bold=True)
styleObj=style(font=boldf)'''
# xu ly du lieu 1
for i in range(2,n+2):
    #sheet.cell(row=1,colum=i).styleObj
    sheet.cell(row=1, column=i).value=i-1
    #sheet.cell(row=i,colum=1).styleObj
    sheet.cell(row=i, column=1).value=i-1
# xu ly du lieu 2
j=2
for r in range (2,n+2):
    for i in range(2,n+2):
        sheet.cell(row=r, column=i).value=sheet.cell(row=1,column=i).value*sheet.cell(row=j,column=1).value
    j=j+1
wb.save('text1.xlsx')
