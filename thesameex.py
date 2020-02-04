import openpyxl
# nhập dữ liệu đầu vào
print('nhập tên file excel chứa sheet cần so sánh: ')
exname=input()
exname=exname+'.xlsx'
print('nhập tên sheet mẫu')
sname=input()
if sname[0].islower:
    sname.capitalize()
print('Mời bạn nhập kí tự đánh dấu sheet khác với sheet gốc: ')
sym=input()
wb=openpyxl.load_workbook(exname)
sheet1=wb.get_sheet_by_name(sname)
# xử lý 1
while True:
    print('Nhập tên sheet cần so sánh (enter để thoát): ')
    saname=input()
    if saname=='':
        wb.save('file đã lọc '+exname)
        break
    if saname[0].islower:
        saname.capitalize()
    sheet2=wb.get_sheet_by_name(saname)
    t=0
    for r in range(1,9):
        if t==1:
            break
        else:
            for c in range(1,9):
                sample=sheet2.cell(row=r,column=c).value
                if sample != sheet1.cell(row=r,column=c).value:
                    print('Khác nhau')
                    sheet2.title=saname+sym
                    t=1
                    break
print('Đã hoàn tất')